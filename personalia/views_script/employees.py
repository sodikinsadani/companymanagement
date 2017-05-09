from ..models import Person,Action,Employee,Leader,ErrorUpload
from ..forms import fEmployee,fPerson,fUplEmp
from django.db import transaction,IntegrityError
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook

def GetEmloyees():
    employee = Employee.objects.all()
    employee = {'employee':employee,
        'action_id' : '2',
    }
    return employee

def editEmp(request,params):
    p = params
    employee = get_object_or_404(Employee,pk=p['params'])
    msg = {}
    form = (fPerson(instance=employee.person),fEmployee(instance=employee))
    if request.method == 'POST':
        cekform1 = fPerson(request.POST or None, instance=employee.person)
        cekform2 = fEmployee(request.POST or None, instance=employee)
        if cekform1.is_valid() and cekform2.is_valid():
            try:
                with transaction.atomic():
                    person = cekform1.save()
                    cekform2.save()
                    form = (cekform1,cekform2)
                    msg = 'Sukses mengubah data %s dengan ID %s'%(person.name,person.id)
            except IntegrityError:
                handle_exception()

    context = {
        'msg':msg,'form':form,'button':('Ubah',),'action':p['action'],
    }
    return context

def newEmp(request,params):
    msg = {}
    form = (fPerson(),fEmployee())
    if request.method == 'POST':
        cekform1 = fPerson(request.POST)
        cekform2 = fEmployee(request.POST)
        if cekform1.is_valid() and cekform2.is_valid():
            try:
                with transaction.atomic():
                    person = cekform1.save()
                    employee = cekform2.save(commit=False)
                    employee.person = person
                    employee.save()
                    msg = 'Berhasil menambahkan %s dengan ID %s'%(person.name,person.id)
            except IntegrityError:
                form = (cekform1,cekform2)
                msg = 'Gagal menambah data karyawan'
    context = {
        'msg':msg,'form':form,'button':('Simpan',),'action':params['action'],
    }
    return context

def handle_upload_file(f):
    wb = load_workbook(f)
    ws = wb['data']

    rows = 5
    total_data = 0
    error_data = 0
    maxrows = ws.max_row
    for d in range(rows,maxrows):
        rows += 1;
        name = ws.cell(row=rows,column=2).value
        gender = ws.cell(row=rows,column=5).value
        grade = ws.cell(row=rows,column=13).value
        status_active = ws.cell(row=rows,column=15).value
        leader = ws.cell(row=rows,column=16).value
        leader = Leader.objects.get(pk=leader)

        birthplace = ws.cell(row=rows,column=3).value
        birth = ws.cell(row=rows,column=4).value
        address = ws.cell(row=rows,column=6).value
        status = ws.cell(row=rows,column=7).value
        school = ws.cell(row=rows,column=8).value
        graduate = ws.cell(row=rows,column=9).value
        mobilephone = ws.cell(row=rows,column=10).value
        bbm = ws.cell(row=rows,column=11).value
        email = ws.cell(row=rows,column=12).value
        date_register = ws.cell(row=rows,column=14).value
        description = ws.cell(row=rows,column=17).value

        try:
            p = Person(name=name,gender=gender)
            p.save()
            employee = Employee(person=p,grade=grade,status_active=status_active,
            leader=leader)
            employee.save()
            total_data += 1
        except IntegrityError:
            errorupload = ErrorUpload(name=name)
            errorupload.save()
            error_data += 1
    msg = '%s data valid, %s data invalid' % (total_data, error_data)
    return msg


def uplEmp(request,params):
    msg = {}
    form = (fUplEmp(),)
    if request.method == 'POST':
        cekformUp = fUplEmp(request.POST, request.FILES)
        if cekformUp.is_valid():
            msg = handle_upload_file(request.FILES['file_upload'])
    context = {
        'msg':msg,'form':form,'button':('Upload',),'action':params['action'],
    }
    return context


actChoice = {'+New':newEmp,'edit':editEmp,'upload':uplEmp}
def GetContext(request,params):
    action_name = params['action'].action_name
    context = actChoice[action_name](request,params)
    return context
