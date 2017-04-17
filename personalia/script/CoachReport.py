from openpyxl import Workbook,load_workbook
from personalia.models import Employee,Coaching
from datetime import datetime as dt
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import Alignment,Font,colors
from django.shortcuts import get_object_or_404

def reportIndividu(response,params):
    p = params
    file = open('personalia/template_xlsx/histCoach.xlsx', 'rb')
    wb = load_workbook(filename = file)
    ws = wb['data']

    employee = get_object_or_404(Employee,pk=p['params'])
    coachHist = Coaching.objects.filter(employee=p['params'])

    rows = 1
    rows += 1;ws.cell(row=rows,column=1,value=params['report_name'].upper())
    date_now = dt.today()
    date_now = date_now.strftime("%d %B %Y")
    rows += 1;ws.cell(row=rows,column=1,value=date_now)
    rows += 1

    #header
    rows += 1;ws.cell(row=rows,column=3,value=': %s' % employee.person.name.title())
    rows += 1;ws.cell(row=rows,column=3,value=': %s' % employee.get_grade_display())
    rows += 1;ws.cell(row=rows,column=3,value=': %s' % employee.leader.name.title())
    rows += 2
    #detail
    for hist in coachHist:
        col = 0
        rows += 1
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')
        col += 1;ws.cell(row=rows,column=col,value=rows-9).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=hist.course.title())
        col += 1;ws.cell(row=rows,column=col,value=hist.date_coaching).number_format = 'dd mmm yyy'
        col += 1;ws.cell(row=rows,column=col,value=hist.description)

    wb.save(response)
    return response
