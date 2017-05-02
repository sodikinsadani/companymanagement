from openpyxl import Workbook,load_workbook
from ..models import Employee
from datetime import datetime as dt
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import Alignment,Font,colors
from django.contrib.auth.models import User

def ConstructReport(response,params):
    report_name = params['reports'].report_name
    file = open('personalia/template_xlsx/tmpEmp.xlsx', 'rb')
    wb = load_workbook(filename = file)
    ws = wb['data']
    wc = wb.create_sheet('chart')
    wc = wb['chart']

    employee = Employee.objects.all().order_by('leader__leader_id','person__gender',
        'grade','status_active')
    rows = 1
    rows += 1;ws.cell(row=rows,column=1,value=report_name.upper())
    date_now = dt.today()
    date_now = date_now.strftime("%d %B %Y")
    rows += 1;ws.cell(row=rows,column=1,value=date_now)
    rows += 2

    for emp in employee:
        col = 0
        rows += 1
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')
        col += 1;ws.cell(row=rows,column=col,value=rows-5).alignment = ac
        #ws.cell(row=rows,column=col).font = ft
        col += 1;ws.cell(row=rows,column=col,value=emp.person.name.upper())
        col += 1;ws.cell(row=rows,column=col,value=emp.person.birthplace.title())
        col += 1;ws.cell(row=rows,column=col,value=emp.person.birth).number_format = 'dd mmm yyy'
        col += 1;ws.cell(row=rows,column=col,value=emp.person.get_gender_display())
        col += 1;ws.cell(row=rows,column=col,value=emp.person.address.title())
        col += 1;ws.cell(row=rows,column=col,value=emp.person.get_status_display()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=emp.person.school.title())
        col += 1;ws.cell(row=rows,column=col,value=emp.person.get_graduate_display()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=emp.person.mobilephone)
        col += 1;ws.cell(row=rows,column=col,value=emp.person.bbm.upper())
        col += 1;ws.cell(row=rows,column=col,value=emp.person.email.lower())
        col += 1;ws.cell(row=rows,column=col,value=emp.get_grade_display()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=emp.date_register).number_format = 'dd mmm yyy'
        col += 1;ws.cell(row=rows,column=col,value=emp.get_status_active_display()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=emp.leader.name.title()).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=emp.description)

    edict = {
        '0':{'0':'WB','L':0,'P':0},
        '1':{'1':'Pra A1','L':0,'P':0},
        '2':{'2':'A1_1','L':0,'P':0},
        '3':{'3':'A1_2','L':0,'P':0},
        '4':{'4':'A1_3','L':0,'P':0},
        '5':{'5':'A2_A','L':0,'P':0},
        '6':{'6':'A2_B','L':0,'P':0},
    }
    erall = Employee.emp_report.Emp("'a' = 'a'")
    for e in erall:
        edict[e[0]][e[1]] = e[2]

    eKeys = edict.keys()
    eKeys.sort()
    wc.append(('Level', 'Laki-laki', 'Perempuan'))
    for e in eKeys:
        wc.append((edict[e][e],edict[e]['L'],edict[e]['P']))

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Seluruh Karyawan"
    chart1.y_axis.title = 'Jumlah'
    chart1.x_axis.title = 'Level Karir'

    data = Reference(wc, min_col=2, min_row=1, max_row=7, max_col=3)
    cats = Reference(wc, min_col=1, min_row=2, max_row=7)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    wc.add_chart(chart1, "A10")

    wb.save(response)
    return response
