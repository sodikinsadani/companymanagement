from openpyxl import Workbook,load_workbook
from ..models import Employee
from datetime import datetime as dt
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import Alignment,Font,colors
from django.contrib.auth.models import User

def GetEmpProg():
    from django.db import connection as con
    with con.cursor() as cursor:
        cursor.execute('''
            select l.manager,l.leader_id,l.name,p0,p1,p2,p3,p4,p5,p6
                ,a0,a1,a2,a3,a4,a5,a6,bk1,bk2,bk3
            from personalia_leader l
            left join (
                select
                      leader_id
                      ,count(decode(e.grade,0,decode(p.gender,'L','0')))||'/'||count(decode(e.grade,0,
                      decode(p.gender,'P','0')))||'/'||count(decode(e.grade,0,'0')) as p0
                      ,count(decode(e.grade,1,decode(p.gender,'L','1')))||'/'||count(decode(e.grade,1,
                      decode(p.gender,'P','1')))||'/'||count(decode(e.grade,1,'1')) as p1
                      ,count(decode(e.grade,2,decode(p.gender,'L','2')))||'/'||count(decode(e.grade,2,
                      decode(p.gender,'P','2')))||'/'||count(decode(e.grade,2,'2')) as p2
                      ,count(decode(e.grade,3,decode(p.gender,'L','3')))||'/'||count(decode(e.grade,3,
                      decode(p.gender,'P','3')))||'/'||count(decode(e.grade,3,'3')) as p3
                      ,count(decode(e.grade,4,decode(p.gender,'L','4')))||'/'||count(decode(e.grade,4,
                      decode(p.gender,'P','4')))||'/'||count(decode(e.grade,4,'4')) as p4
                      ,count(decode(e.grade,5,decode(p.gender,'L','5')))||'/'||count(decode(e.grade,5,
                      decode(p.gender,'P','5')))||'/'||count(decode(e.grade,5,'5')) as p5
                      ,count(decode(e.grade,6,decode(p.gender,'L','6')))||'/'||count(decode(e.grade,6,
                      decode(p.gender,'P','6')))||'/'||count(decode(e.grade,6,'6')) as p6
                      from personalia_employee e
                      inner join personalia_person p on p.id = e.person_id
                      where status_active = 'pa'
                      group by leader_id
            ) pa on pa.leader_id = l.leader_id
            left join (
                select
                      leader_id
                      ,count(decode(e.grade,0,decode(p.gender,'L','0')))||'/'||count(decode(e.grade,0,
                      decode(p.gender,'P','0')))||'/'||count(decode(e.grade,0,'0')) as a0
                      ,count(decode(e.grade,1,decode(p.gender,'L','1')))||'/'||count(decode(e.grade,1,
                      decode(p.gender,'P','1')))||'/'||count(decode(e.grade,1,'1')) as a1
                      ,count(decode(e.grade,2,decode(p.gender,'L','2')))||'/'||count(decode(e.grade,2,
                      decode(p.gender,'P','2')))||'/'||count(decode(e.grade,2,'2')) as a2
                      ,count(decode(e.grade,3,decode(p.gender,'L','3')))||'/'||count(decode(e.grade,3,
                      decode(p.gender,'P','3')))||'/'||count(decode(e.grade,3,'3')) as a3
                      ,count(decode(e.grade,4,decode(p.gender,'L','4')))||'/'||count(decode(e.grade,4,
                      decode(p.gender,'P','4')))||'/'||count(decode(e.grade,4,'4')) as a4
                      ,count(decode(e.grade,5,decode(p.gender,'L','5')))||'/'||count(decode(e.grade,5,
                      decode(p.gender,'P','5')))||'/'||count(decode(e.grade,5,'5')) as a5
                      ,count(decode(e.grade,6,decode(p.gender,'L','6')))||'/'||count(decode(e.grade,6,
                      decode(p.gender,'P','6')))||'/'||count(decode(e.grade,6,'6')) as a6
                      from personalia_employee e
                      inner join personalia_person p on p.id = e.person_id
                      where status_active = 'ak'
                      group by leader_id
            ) ak on ak.leader_id = l.leader_id
            left join (
                select
                      leader_id
                      ,count(decode(e.status_active,'bk1',decode(p.gender,'L','bk1')))||'/'||count(decode(e.status_active,'bk1',
                      decode(p.gender,'P','bk1')))||'/'||count(decode(e.status_active,'bk1','bk1')) as bk1
                      ,count(decode(e.status_active,'bk2',decode(p.gender,'L','bk2')))||'/'||count(decode(e.status_active,'bk2',
                      decode(p.gender,'P','bk2')))||'/'||count(decode(e.status_active,'bk2','bk2')) as bk2
                      ,count(decode(e.status_active,'bk3',decode(p.gender,'L','bk3')))||'/'||count(decode(e.status_active,'bk3',
                      decode(p.gender,'P','bk3')))||'/'||count(decode(e.status_active,'bk3','bk3')) as bk3
                      from personalia_employee e
                      inner join personalia_person p on p.id = e.person_id
                      where status_active in ('bk1','bk2','bk3')
                      group by leader_id
            ) bk on bk.leader_id = l.leader_id
            order by l.manager,l.name
        ''')
        result_list = cursor.fetchall()
    return result_list

def ConstructReport(response,params):
    report_name = params['reports'].report_name
    file = open('personalia/template_xlsx/tmpPerMon.xlsx', 'rb')
    wb = load_workbook(filename = file)

    result_list = GetEmpProg()

    sDict = {
    'ahm':('ahmad',),'ksm':('kusmawan',),'mgd':('migud',),
    'sdk':('sodikin',),'sdr':('sodirun',)
    }


    manager = ''
    for r in result_list:
        ft = Font(color=colors.RED)
        ac = Alignment(horizontal='center')

        if manager != r[0] :
            manager = r[0]
            ws = wb[sDict[manager][0]]
            rows = 7
            no = 0

            ws.cell(row=2,column=1,value=report_name.upper())
            date_now = dt.today()
            date_now = date_now.strftime("%d %B %Y")
            ws.cell(row=3,column=1,value=date_now)

        no += 1
        col = 0
        rows += 1
        col += 1;ws.cell(row=rows,column=col,value=no).alignment = ac
        col += 1;ws.cell(row=rows,column=col,value=r[2].upper())

        for c in range(3,20):
            col += 1;ws.cell(row=rows,column=col,value=r[c] or '0/0/0').alignment = ac

    wb.save(response)
    return response
